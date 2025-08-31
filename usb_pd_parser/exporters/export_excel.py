import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Regex to remove illegal characters
ILLEGAL_CHARACTERS_RE = re.compile(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]')

def clean_text(value):
    """Convert to string and strip illegal Excel characters"""
    if value is None:
        return ""
    if isinstance(value, list):
        value = " ".join(str(v) for v in value)
    text = str(value)
    return ILLEGAL_CHARACTERS_RE.sub("", text)

def export_to_excel(data, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sections"

    if not data:
        wb.save(out_path)
        return

    # --- Sections sheet ---
    headers = list(data[0].keys())
    ws.append([clean_text(h) for h in headers])

    for r in data:
        row = [clean_text(r.get(h, "")) for h in headers]
        ws.append(row)

    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # --- TOC sheet ---
    toc = wb.create_sheet("TOC")
    toc.append(["Section ID", "Title"])

    for r in data:
        toc.append([r.get("section_id", ""), r.get("title", "")])

    # Style TOC headers
    for cell in toc[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="9BBB59")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(out_path)
